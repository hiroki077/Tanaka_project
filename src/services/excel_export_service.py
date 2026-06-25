from __future__ import annotations
import hashlib
import logging
import re
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Callable, Protocol

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import (
    AnchorMarker, TwoCellAnchor,
)
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils.cell import (
    coordinate_from_string, column_index_from_string, get_column_letter,
)
from PIL import Image as PILImage, ImageDraw, ImageFont

from ..config import (
    DEFAULT_PHOTO_WIDTH_COLS,
    DEFAULT_PHOTO_HEIGHT_ROWS,
    DEFAULT_PHOTO_OFFSET_ROWS_UP,
    PLACEHOLDER_PATTERN,
)
from ..db import (
    Database,
    Employee,
    EmploymentStatus,
    MappingOverrideRepository,
    strip_name_prefix_marks,
)
from ..services.employee_service import EmployeeService
from ..services.photo_service import PhotoService


log = logging.getLogger(__name__)


class MatchMode(str, Enum):
    PLACEHOLDER_ONLY = "placeholder_only"
    NAME_DETECT_ONLY = "name_detect_only"
    HYBRID = "hybrid"


@dataclass
class ExportOptions:
    template_path: Path
    output_path: Path
    target_sheets: list[str] | None = None
    match_mode: MatchMode = MatchMode.HYBRID
    photo_width_cols: int = DEFAULT_PHOTO_WIDTH_COLS
    photo_height_rows: int = DEFAULT_PHOTO_HEIGHT_ROWS
    name_detect_offset_rows_up: int = DEFAULT_PHOTO_OFFSET_ROWS_UP
    keep_existing_images: bool = True

    def template_signature(self) -> str:
        return f"{self.template_path.name}:{_file_hash(self.template_path)[:16]}"


@dataclass
class AmbiguousMatch:
    sheet_name: str
    cell_address: str
    match_key: str
    candidates: list[Employee]


@dataclass
class ExportResult:
    sheets_processed: int = 0
    photos_inserted: int = 0
    skipped_no_photo: list[str] = field(default_factory=list)
    skipped_on_leave: list[str] = field(default_factory=list)
    unmatched: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    output_path: Path | None = None


class AmbiguityResolver(Protocol):
    """同名解決のための関数プロトコル。

    呼び出し元（GUI/CLI/テスト）が同名候補の選択方法を注入する。
    `None` を返した場合は「該当なし／スキップ」として扱う。
    """

    def __call__(self, ambiguous: AmbiguousMatch) -> Employee | None: ...


def first_candidate_resolver(ambiguous: AmbiguousMatch) -> Employee | None:
    return ambiguous.candidates[0] if ambiguous.candidates else None


def skip_resolver(ambiguous: AmbiguousMatch) -> Employee | None:
    return None


# 入社年っぽいセル（4桁年。M2017 / 2019C / 2024 など前後に英字 1〜2 文字を許容）
_JOIN_YEAR_RE = re.compile(
    r"^[\sA-Za-z　Ａ-Ｚａ-ｚ]{0,3}\d{4}[\sA-Za-z　Ａ-Ｚａ-ｚ]{0,3}$"
)

# 氏名セルから右側に何セル分・上下何行分まで「入社年セル」を探索するか
_YEAR_SEARCH_COLS = 8
_YEAR_SEARCH_ROWS_ABOVE = 2
_YEAR_SEARCH_ROWS_BELOW = 1


def _clear_fill(cell) -> None:
    """セルの背景塗りつぶしを透明（fill_type=None）にする。"""
    cell.fill = PatternFill(fill_type=None)


# テキスト画像化用のフォント候補（プラットフォーム横断）
_FONT_PATHS = (
    "/System/Library/Fonts/Helvetica.ttc",
    "/System/Library/Fonts/Supplemental/Arial.ttf",
    "C:\\Windows\\Fonts\\arial.ttf",
    "C:\\Windows\\Fonts\\meiryo.ttc",
    "C:\\Windows\\Fonts\\msgothic.ttc",
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
)


def _load_font(size_px: int):
    for path in _FONT_PATHS:
        try:
            return ImageFont.truetype(path, size_px)
        except (OSError, IOError):
            continue
    return ImageFont.load_default()


def _cell_text_color(cell) -> tuple[int, int, int]:
    try:
        c = cell.font.color
        if c and getattr(c, "type", None) == "rgb" and c.rgb:
            rgb = str(c.rgb)
            if len(rgb) >= 6:
                hex6 = rgb[-6:]
                return (int(hex6[0:2], 16), int(hex6[2:4], 16), int(hex6[4:6], 16))
    except (AttributeError, ValueError, TypeError):
        pass
    return (0, 0, 0)


def _cell_font_size_pt(cell, default: float = 11.0) -> float:
    try:
        if cell.font and cell.font.sz:
            return float(cell.font.sz)
    except (AttributeError, TypeError):
        pass
    return default


def _cell_horizontal_align(cell) -> str:
    try:
        if cell.alignment and cell.alignment.horizontal:
            return cell.alignment.horizontal
    except AttributeError:
        pass
    return "center"


def _render_text_png(
    text: str,
    width_px: int,
    height_px: int,
    font_size_pt: float,
    color: tuple[int, int, int],
    align: str,
) -> bytes:
    """テキストを透明背景PNGにレンダリングしてバイト列で返す。"""
    import io
    font_size_px = max(6, int(round(font_size_pt * 4 / 3)))
    width_px = max(width_px, 30)
    height_px = max(height_px, 16)

    img = PILImage.new("RGBA", (width_px, height_px), (255, 255, 255, 0))
    draw = ImageDraw.Draw(img)
    font = _load_font(font_size_px)

    bbox = draw.textbbox((0, 0), text, font=font)
    tw = bbox[2] - bbox[0]
    th = bbox[3] - bbox[1]

    pad = 4
    if align == "right":
        x = width_px - tw - pad
    elif align == "left":
        x = pad
    else:
        x = (width_px - tw) // 2
    y = (height_px - th) // 2 - bbox[1]

    draw.text((x, y), text, fill=(*color, 255), font=font)

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def _replace_year_cell_with_image(
    ws, anchor_cell, temp_files: list[Path]
) -> None:
    """年セルの値をPNG画像に置き換えてセル値を空にする。

    Excel ではセルに値があると、左隣セルからオーバーフローしてきたふりがなが
    そのセル境界で切られる。値を消して空セル化することでオーバーフローを許す。
    年の見た目は同位置に浮かべた画像で維持する。
    """
    value = anchor_cell.value
    if value is None:
        return
    text = str(value).strip()
    if not text:
        return

    # 結合範囲を解決
    target_row = anchor_cell.row
    target_col = anchor_cell.column
    min_row, min_col = target_row, target_col
    max_row, max_col = target_row, target_col
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= target_row <= mr.max_row and mr.min_col <= target_col <= mr.max_col:
            min_row, min_col = mr.min_row, mr.min_col
            max_row, max_col = mr.max_row, mr.max_col
            break

    anchor = ws.cell(row=min_row, column=min_col)

    # 結合範囲合計の幅・高さ（px）
    width_px = 0
    for c in range(min_col, max_col + 1):
        cw = ws.column_dimensions[get_column_letter(c)].width or 8.43
        width_px += int(cw * 7)
    height_px = 0
    for r in range(min_row, max_row + 1):
        rh = ws.row_dimensions[r].height or 15
        height_px += int(rh * 4 / 3)

    font_size_pt = _cell_font_size_pt(anchor)
    color = _cell_text_color(anchor)
    align = _cell_horizontal_align(anchor)

    png_bytes = _render_text_png(
        text, width_px, height_px, font_size_pt, color, align
    )

    import tempfile
    tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    tmp.write(png_bytes)
    tmp.close()
    temp_files.append(Path(tmp.name))

    xl_img = XLImage(tmp.name)
    xl_img.width = width_px
    xl_img.height = height_px

    image_anchor = TwoCellAnchor(
        editAs="oneCell",
        _from=AnchorMarker(col=min_col - 1, colOff=0, row=min_row - 1, rowOff=0),
        to=AnchorMarker(col=max_col, colOff=0, row=max_row, rowOff=0),
    )
    xl_img.anchor = image_anchor
    ws.add_image(xl_img)

    anchor.value = None
    _clear_fill(anchor)


def _replace_adjacent_year_cells_with_images(
    ws, name_cell_coord: str, temp_files: list[Path]
) -> None:
    """氏名セルの右隣付近にある入社年セルを画像化して退避する。

    対象範囲: 上下数行 × 右数列。年っぽい値を持つセルを全て画像化する。
    結合セル内ならその左上アンカーで処理する。
    """
    col_letter, row_num = coordinate_from_string(name_cell_coord)
    col_idx = column_index_from_string(col_letter)

    merged_ranges = list(ws.merged_cells.ranges)
    processed: set[tuple[int, int]] = set()

    for dr in range(-_YEAR_SEARCH_ROWS_ABOVE, _YEAR_SEARCH_ROWS_BELOW + 1):
        for dc in range(1, _YEAR_SEARCH_COLS + 1):
            r = row_num + dr
            c = col_idx + dc
            if r < 1 or c < 1:
                continue
            anchor_r, anchor_c = r, c
            for mr in merged_ranges:
                if mr.min_row <= r <= mr.max_row and mr.min_col <= c <= mr.max_col:
                    anchor_r, anchor_c = mr.min_row, mr.min_col
                    break
            if (anchor_r, anchor_c) in processed:
                continue
            anchor = ws.cell(row=anchor_r, column=anchor_c)
            v = anchor.value
            if v is None:
                continue
            s = str(v).strip()
            if not _JOIN_YEAR_RE.match(s):
                continue
            _replace_year_cell_with_image(ws, anchor, temp_files)
            processed.add((anchor_r, anchor_c))


def _enable_shrink_to_fit(cell) -> None:
    """既存の Alignment 設定を維持したまま shrink_to_fit のみ True にする。"""
    a = cell.alignment
    cell.alignment = Alignment(
        horizontal=a.horizontal,
        vertical=a.vertical,
        text_rotation=a.text_rotation,
        wrap_text=False,  # shrink_to_fit と wrap_text は排他
        shrink_to_fit=True,
        indent=a.indent,
        relativeIndent=a.relativeIndent,
        justifyLastLine=a.justifyLastLine,
        readingOrder=a.readingOrder,
    )


def _file_hash(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


_PLACEHOLDER_RE = re.compile(PLACEHOLDER_PATTERN)
_DETAIL_RE = re.compile(r"(?P<key>[^#]+?)(?:#id=(?P<id>\d+))?$")

# 写真をセル枠の内側に数pxだけ縮めて配置するためのオフセット（EMU 単位）。
# 1 px ≒ 9525 EMU (96DPI 換算)。0 だと写真縁がセル境界に一致して青枠を覆い隠す。
_PHOTO_INSET_EMU = 28575  # 約 3 px


class ExcelExportService:
    def __init__(
        self,
        db: Database,
        employee_service: EmployeeService,
        photo_service: PhotoService,
        ambiguity_resolver: AmbiguityResolver = first_candidate_resolver,
    ):
        self.db = db
        self.employees = employee_service
        self.photos = photo_service
        self.override_repo = MappingOverrideRepository(db)
        self.ambiguity_resolver = ambiguity_resolver

    def export(self, opts: ExportOptions) -> ExportResult:
        if not opts.template_path.is_file():
            raise FileNotFoundError(f"テンプレートが見つかりません: {opts.template_path}")
        wb = load_workbook(opts.template_path)

        result = ExportResult()
        template_sig = opts.template_signature()
        self._temp_files: list[Path] = []

        target_sheet_names = opts.target_sheets or wb.sheetnames

        try:
            for sheet_name in target_sheet_names:
                if sheet_name not in wb.sheetnames:
                    result.warnings.append(f"シート '{sheet_name}' が見つかりません")
                    continue
                ws = wb[sheet_name]
                self._process_sheet(ws, sheet_name, opts, template_sig, result)
                result.sheets_processed += 1

            opts.output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(opts.output_path)
        finally:
            for tf in self._temp_files:
                try:
                    tf.unlink(missing_ok=True)
                except Exception:
                    pass
            self._temp_files = []

        result.output_path = opts.output_path
        return result

    def _process_sheet(
        self,
        ws,
        sheet_name: str,
        opts: ExportOptions,
        template_sig: str,
        result: ExportResult,
    ) -> None:
        # 1) プレースホルダー走査
        if opts.match_mode in (MatchMode.PLACEHOLDER_ONLY, MatchMode.HYBRID):
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    text = str(cell.value)
                    m = _PLACEHOLDER_RE.search(text)
                    if not m:
                        continue
                    raw_key = m.group(1).strip()
                    emp = self._resolve_detail_key(raw_key)
                    if isinstance(emp, list):
                        emp = self._resolve_ambiguity(
                            template_sig, sheet_name, cell.coordinate,
                            raw_key, emp, result
                        )
                    if emp is None:
                        result.unmatched.append(
                            f"[{sheet_name}!{cell.coordinate}] '{raw_key}' に該当者なし"
                        )
                        cell.value = ""
                        continue
                    if not self._can_insert(emp, sheet_name, cell.coordinate, result):
                        cell.value = ""
                        continue
                    self._insert_photo(
                        ws, cell.coordinate, emp, opts, anchor_at_cell=True
                    )
                    cell.value = ""
                    result.photos_inserted += 1

        # 2) 氏名直接検出（プレースホルダーが既に処理したセルは値クリア済みなのでスキップされる）
        if opts.match_mode in (MatchMode.NAME_DETECT_ONLY, MatchMode.HYBRID):
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    text = str(cell.value).strip()
                    if not text or _PLACEHOLDER_RE.search(text):
                        continue
                    candidates = self.employees.repo.find_by_text(
                        text, only_active=False,
                    )
                    if not candidates:
                        continue

                    # ★ 氏名マッチした時点で見た目クリーンアップを実施。
                    # 写真挿入の成否（休職/写真未登録/同名解決失敗等）に関わらず、
                    # 先頭記号除去・shrink_to_fit・年セル退避は走らせる。
                    cleaned = strip_name_prefix_marks(str(cell.value))
                    if cleaned != str(cell.value):
                        cell.value = cleaned
                    _enable_shrink_to_fit(cell)
                    _replace_adjacent_year_cells_with_images(
                        ws, cell.coordinate, self._temp_files,
                    )

                    active = [e for e in candidates if e.status == EmploymentStatus.ACTIVE.value]
                    if not active:
                        result.skipped_on_leave.append(
                            f"[{sheet_name}!{cell.coordinate}] '{text}' 全員 休職/退職"
                        )
                        continue
                    if len(active) == 1:
                        emp = active[0]
                    else:
                        emp = self._resolve_ambiguity(
                            template_sig, sheet_name, cell.coordinate,
                            text, active, result
                        )
                    if emp is None or not self._can_insert(emp, sheet_name, cell.coordinate, result):
                        continue
                    self._insert_photo(
                        ws, cell.coordinate, emp, opts, anchor_at_cell=False
                    )
                    result.photos_inserted += 1

    def _resolve_detail_key(self, raw_key: str) -> Employee | list[Employee] | None:
        """プレースホルダー内文字列 (`氏名#id=42` 等) を解決。

        - `名前#id=42` : 従業員ID指定
        - `名前`       : match_key 完全一致
        """
        m = _DETAIL_RE.match(raw_key)
        if not m:
            return None
        key = m.group("key").strip()
        explicit_id = m.group("id")

        if explicit_id:
            emp = self.employees.get(int(explicit_id))
            if emp and emp.status == EmploymentStatus.ACTIVE.value:
                return emp
            return None

        # プレースホルダーは「先頭一致での1キー指定」前提なので
        # まず match_key、ヒットしなければ fuzzy（text）検索にフォールバック
        candidates = self.employees.repo.find_by_match_key(key, only_active=True)
        if not candidates:
            candidates = self.employees.repo.find_by_text(key, only_active=True)
        if not candidates:
            return None
        if len(candidates) == 1:
            return candidates[0]
        return candidates

    def _resolve_ambiguity(
        self,
        template_sig: str,
        sheet_name: str,
        cell_address: str,
        match_key: str,
        candidates: list[Employee],
        result: ExportResult,
    ) -> Employee | None:
        # 1) 学習済みオーバーライドを優先
        override = self.override_repo.find(template_sig, sheet_name, cell_address)
        if override:
            for c in candidates:
                if c.id == override.employee_id:
                    return c

        # 2) リゾルバ（GUIなら対話、CLIならデフォルト）に委譲
        chosen = self.ambiguity_resolver(AmbiguousMatch(
            sheet_name=sheet_name,
            cell_address=cell_address,
            match_key=match_key,
            candidates=candidates,
        ))
        if chosen is not None:
            self.override_repo.save(
                template_sig, sheet_name, cell_address, match_key, chosen.id
            )
        else:
            result.warnings.append(
                f"[{sheet_name}!{cell_address}] 同名 {len(candidates)}件 → スキップ"
            )
        return chosen

    def _can_insert(
        self,
        emp: Employee,
        sheet_name: str,
        cell_address: str,
        result: ExportResult,
    ) -> bool:
        if emp.status != EmploymentStatus.ACTIVE.value:
            result.skipped_on_leave.append(
                f"[{sheet_name}!{cell_address}] {emp.name} は {emp.status}"
            )
            return False
        photo = self.photos.resolve(emp.photo_path)
        if photo is None:
            result.skipped_no_photo.append(
                f"[{sheet_name}!{cell_address}] {emp.name} の写真が未登録"
            )
            return False
        return True

    def _insert_photo(
        self,
        ws,
        cell_address: str,
        emp: Employee,
        opts: ExportOptions,
        anchor_at_cell: bool,
    ) -> None:
        photo_path = self.photos.resolve(emp.photo_path)
        if photo_path is None:
            return

        col_letter, row_num = coordinate_from_string(cell_address)
        col_idx = column_index_from_string(col_letter)  # 1-based

        if anchor_at_cell:
            # プレースホルダー: そのセル左上を起点
            top_left_col_0 = col_idx - 1
            top_left_row_0 = row_num - 1
        else:
            # 氏名検出: そのセルの上方 N行 が写真の起点
            top_left_col_0 = col_idx - 1
            top_left_row_0 = max(0, row_num - 1 - opts.name_detect_offset_rows_up)

        bottom_right_col_0 = top_left_col_0 + opts.photo_width_cols
        bottom_right_row_0 = top_left_row_0 + opts.photo_height_rows

        img = XLImage(str(photo_path))
        # ファイルハンドルを早期にクローズし、保存時の汚染を避ける（Pillowで再読込）
        try:
            with PILImage.open(photo_path) as pil:
                img.width, img.height = pil.size
        except Exception:
            pass

        # 写真をセル枠の内側に数px分だけ縮める。
        # セル境界ぴったりだとテンプレの青枠が写真下に隠れて見えなくなるため、
        # 4辺すべてに内側オフセットを入れて枠線を露出させる。
        anchor = TwoCellAnchor(
            editAs="oneCell",
            _from=AnchorMarker(col=top_left_col_0, colOff=_PHOTO_INSET_EMU,
                               row=top_left_row_0, rowOff=_PHOTO_INSET_EMU),
            to=AnchorMarker(col=bottom_right_col_0, colOff=-_PHOTO_INSET_EMU,
                            row=bottom_right_row_0, rowOff=-_PHOTO_INSET_EMU),
        )
        img.anchor = anchor
        ws.add_image(img)
