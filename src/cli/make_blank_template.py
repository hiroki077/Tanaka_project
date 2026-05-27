"""完成版の体制表 xlsx から「写真だけ削除した空テンプレート」を作成する。

完成版 (例: 2605顔写真体制.xlsx) はそのままテンプレとして使うと写真が
二重貼りされるため、写真画像のみ除去した版を作成する。テキストレイアウト
（カナ氏名、漢字氏名、入社年、各種ラベル）はそのまま保持する。

このスクリプトの出力ファイルを月次テンプレートとして使い、本アプリの
「Excel出力」で写真を流し込めば、完成版と同じレイアウトの新しい体制表が
得られる。

実行例:
    python3 -m src.cli.make_blank_template file/2605顔写真体制.xlsx file/template_blank.xlsx
"""
from __future__ import annotations
import argparse
import sys
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")

from openpyxl import load_workbook


# テンプレ化に残したい画像形式は無し（全画像削除）。
# emf/wmf 等の装飾画像も削除対象に含めるかは要件次第。
# 今回は **写真（jpeg/png）のみ削除** し、装飾（emf/wmf 等）は残す方針。
PHOTO_FORMATS = {"jpeg", "jpg", "png"}


def strip_photos_in_sheet(ws) -> int:
    """シート内の写真画像を削除し、削除枚数を返す。"""
    kept = []
    removed = 0
    for img in ws._images:
        fmt = (getattr(img, "format", "") or "").lower()
        if fmt in PHOTO_FORMATS:
            removed += 1
        else:
            kept.append(img)
    ws._images = kept
    return removed


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source", type=Path, help="元のxlsx（完成版）")
    parser.add_argument("dest", type=Path, help="出力先xlsx（空テンプレート）")
    args = parser.parse_args()

    if not args.source.is_file():
        print(f"[error] 元ファイルが見つかりません: {args.source}", file=sys.stderr)
        return 1

    print(f"📂 読み込み: {args.source}")
    wb = load_workbook(args.source)

    total_removed = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        n = strip_photos_in_sheet(ws)
        if n:
            print(f"  - {sn}: 写真 {n}枚を削除")
            total_removed += n

    args.dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.dest)
    print(f"\n✅ 空テンプレートを書き出しました: {args.dest}")
    print(f"   削除した写真総数: {total_removed}枚")
    print(f"\nこのファイルを「Excel出力」のテンプレートに指定すれば、")
    print(f"DBの写真が同じレイアウトで貼られた完成版が生成されます。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
