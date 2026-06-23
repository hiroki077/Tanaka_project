"""202605.xlsx (テキストマスター一覧) から、支店別の顔写真レイアウトを自動生成する。

入力: 202605.xlsx 形式の「テキストマスター体制表」
処理: 各支店ブロックをパース → 支店ごとのシートを生成 → DB から写真を貼り込み
出力: 2605顔写真体制.xlsx と同等の「支店別グリッドレイアウト」xlsx

実行例:
    python3 -m src.cli.build_from_master file/202605.xlsx file/output.xlsx
"""
from __future__ import annotations
import argparse
import re
import sys
import warnings
from dataclasses import dataclass, field
from pathlib import Path

warnings.filterwarnings("ignore")

from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from ..config import Settings, DataPaths, PLACEHOLDER_PHOTO_PATH
from ..db import Database, EmploymentStatus
from ..services import EmployeeService, PhotoService


# ============================================================
# 入力（202605.xlsx）の構造定義
# ============================================================
COL_F = 6   # ブロックヘッダ列（"集合XX支店" や "支店長" 等）
COL_G = 7   # 上記の対応値列（支店長の氏名等）
COL_I = 9   # セクション列 (●営業課, ●設計課)
COL_J = 10  # 課名列 (営業1課, 営業2課, 営業3課)

# 名前列と入社年列のペア（総合職）。L,M / N,O / P,Q / ...
TOTAL_PAIRS = [(c, c + 1) for c in range(12, 32, 2)]  # L-AE まで10ペア
JIMU_COLS = [32, 33]      # AF, AG: 実務職・事務
KEIYAKU_COLS = [34, 35]   # AH, AI: 契約社員
HAKEN_COLS = [36, 37, 38] # AJ, AK, AL: 派遣社員
HOLIDAY_COL = 39          # AM: 休職中（写真は不要だが情報として保持）

BRANCH_RE = re.compile(r"^(集合|《本部スタッフ》|技術部|事業推進部)")
EXCLUDE_BRANCH_RE = re.compile(r"^(本部長|本部長・部長|本部長・部長・支店長)")
SECTION_RE = re.compile(r"^[●○]\s*(.+)")
ROLE_MARK_CHARS = "＊*☆★■◆○◯〇▽▼●◎▲△□"
ROLE_MARK_RE = re.compile(rf"^([{ROLE_MARK_CHARS}]+)?(.+)$")

# 名前末尾に付く役職/契約区分マーカー (M=マネージャー、C=契約 など)
TRAILING_MARK_RE = re.compile(r"[MＭCＣ]+$")

# 飾り文字列・非人物セルを除外するためのパターン
NOT_A_PERSON_RE = re.compile(r"^[《【〈\[].*[》】〉\]]$|^#REF!$|^＝|^=|^兼務除く$")


@dataclass
class PersonEntry:
    raw_text: str
    name: str
    marks: str
    year: str | None
    category: str = "総合職"   # 総合職 / 実務職 / 契約 / 派遣


@dataclass
class CourseRow:
    course_name: str   # "営業1課" / "" / "設計課"
    persons: list[PersonEntry]
    # 同 master 行で同時に取り込まれる事務職/契約/派遣（per-課 表示用）
    jimu: list[PersonEntry] = field(default_factory=list)
    keiyaku: list[PersonEntry] = field(default_factory=list)
    haken: list[PersonEntry] = field(default_factory=list)


@dataclass
class BranchData:
    branch_name: str
    top_positions: list[tuple[str, str]]   # (役職, 氏名)
    sections: dict[str, list[CourseRow]]   # "営業課" → [CourseRow, ...]
    jimu: list[PersonEntry]
    keiyaku: list[PersonEntry]
    haken: list[PersonEntry]


_CONCURRENT_RE = re.compile(r"^[\s　]*兼\s*[)）)）]\s*")


def _parse_person(text: str, year, category: str = "総合職") -> PersonEntry | None:
    """セルの文字列から人物データを抽出。人物セルでなければ None を返す。"""
    text = text.strip()
    if not text or NOT_A_PERSON_RE.match(text):
        return None
    # 兼務プレフィックス除去（marks に "兼" として保持）
    extra_marks = ""
    if _CONCURRENT_RE.match(text):
        text = _CONCURRENT_RE.sub("", text)
        extra_marks = "兼"
    # 先頭の役職記号 (☆, ＊, ○ など)
    m = ROLE_MARK_RE.match(text)
    if m and m.group(1):
        marks = m.group(1)
        name = m.group(2).strip()
    else:
        marks = ""
        name = text
    # 末尾の M/C 等（マネージャー、契約等）。マッチング用に除去し marks に追加
    tail = TRAILING_MARK_RE.search(name)
    if tail:
        marks = marks + tail.group(0)
        name = TRAILING_MARK_RE.sub("", name).strip()
    if extra_marks:
        marks = extra_marks + marks
    # プレースホルダー名（未確定枠）は人物として扱わない
    if name in ("新人", "新 人", "（新人）", "(新人)"):
        return None
    return PersonEntry(
        raw_text=text, name=name, marks=marks,
        year=(str(year).strip() if year is not None else None),
        category=category,
    )


def parse_master(path: Path) -> list[BranchData]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    branches: list[BranchData] = []
    current: BranchData | None = None
    current_section: str | None = None

    def flush():
        nonlocal current
        if current:
            branches.append(current)
        current = None

    last_course_row: CourseRow | None = None
    section_closed = False   # AD列 '兼務除く' 検出後 True、新しい ● セクションでリセット

    for row in range(1, ws.max_row + 1):
        f = ws.cell(row=row, column=COL_F).value
        g = ws.cell(row=row, column=COL_G).value
        i = ws.cell(row=row, column=COL_I).value
        j = ws.cell(row=row, column=COL_J).value

        if f and isinstance(f, str):
            text = f.strip()
            if EXCLUDE_BRANCH_RE.match(text):
                continue
            if BRANCH_RE.match(text):
                flush()
                current = BranchData(
                    branch_name=text,
                    top_positions=[], sections={},
                    jimu=[], keiyaku=[], haken=[],
                )
                current_section = None
                last_course_row = None
                section_closed = False

        if not current:
            continue

        if f and isinstance(f, str) and g:
            text = f.strip()
            if not BRANCH_RE.match(text):
                current.top_positions.append((text, str(g).strip()))

        if i and isinstance(i, str):
            m = SECTION_RE.match(i.strip())
            if m:
                current_section = m.group(1).strip()
                current.sections.setdefault(current_section, [])
                section_closed = False   # 新しいセクション開始
                last_course_row = None

        # 課行（総合職）
        new_course_row: CourseRow | None = None
        if current_section is not None and not section_closed:
            course_name = ""
            if j and isinstance(j, str):
                course_name = j.strip()
            persons = []
            for name_col, year_col in TOTAL_PAIRS:
                n = ws.cell(row=row, column=name_col).value
                y = ws.cell(row=row, column=year_col).value
                if n and isinstance(n, str) and n.strip():
                    p = _parse_person(str(n), y, "総合職")
                    if p:
                        persons.append(p)
            if persons:
                new_course_row = CourseRow(course_name=course_name, persons=persons)
                current.sections[current_section].append(new_course_row)
                last_course_row = new_course_row

        # 実務職・契約・派遣 → branch 単位 + 同 master 行は直近 CourseRow にも紐付け
        # section_closed の場合は branch にも紐付けない（GM 候補等の混入を防ぐ）
        if not section_closed:
            cr_attach = new_course_row or last_course_row
            for col, branch_bucket, cr_attr, cat in [
                *[(c, current.jimu, "jimu", "実務職") for c in JIMU_COLS],
                *[(c, current.keiyaku, "keiyaku", "契約") for c in KEIYAKU_COLS],
                *[(c, current.haken, "haken", "派遣") for c in HAKEN_COLS],
            ]:
                v = ws.cell(row=row, column=col).value
                if v and isinstance(v, str) and v.strip():
                    p = _parse_person(v, None, cat)
                    if p:
                        branch_bucket.append(p)
                        if cr_attach is not None:
                            getattr(cr_attach, cr_attr).append(p)

        # AD列(col 30) の '兼務除く' でセクション終了を検出
        ad_val = ws.cell(row=row, column=30).value
        if isinstance(ad_val, str) and "兼務除く" in ad_val:
            section_closed = True

    flush()
    return branches


# ============================================================
# 出力レイアウト （202605_顔写真版_修正依頼0605.xlsx 仕様 準拠）
# ============================================================
FONT_NAME = "Meiryo"  # 役職12/氏名11/フリガナ年次10/タイトル28、太字使い分け

PERSON_COL_W = 6      # 1人スロット = 5列(content) + 1列(gap)
PHOTO_ROWS = 8        # 写真の高さ（行数）
LABEL_ROWS = 3        # ラベル3行: カナ / 漢字 / 年
LEADER_LABEL_ROWS = 1
PERSON_TOTAL_H = LEADER_LABEL_ROWS + PHOTO_ROWS + LABEL_ROWS + 1

# サブセクション/役職ラベル用の縦方向間隔
SECTION_HEADER_ROWS = 1
SUBSECTION_ROWS = 1
ROLE_LABEL_ROWS = 1
SPACER_ROWS = 1

# 列構成: col A = 左マージン, col B 以降 = 等幅, person 1 は col C 開始
SHEET_START_COL = 3   # col C
MARGIN_COL_WIDTH = 2.625
UNIFORM_COL_WIDTH = 3.0   # 依頼ファイル準拠だと 13.0 で広すぎ → 画面表示用に縮小
# タイトル/印刷範囲。派遣社員等を含む横長行で最大 10 人配置されるケース
# (slot 9 = col 57) があるためそれをカバーする。
DATA_COL_END = 60

# 行高
TITLE_ROW_HEIGHT = 50.25
TOP_HEADER_HEIGHT = 20.25     # 支店長/支店付 ラベル行
SECTION_HEADER_HEIGHT = 19.5  # 「営 業 課」「設 計 課」
SUBSECTION_HEIGHT = 19.5      # 「営業1課」「営業2課」
ROLE_LABEL_HEIGHT = 19.5      # 「課長」「課長代務」
SPACER_HEIGHT = 6.0           # ラベル間のスペーサー行
# 写真比率 3:4 (col_width 3.0 × 5列 = 105px / row_height 13.13 × 8行 = 140px)
PHOTO_ROW_HEIGHT = 13.13
LABEL_ROW_HEIGHT = 14.1       # カナ/漢字/年

# カラーパレット（依頼ファイル準拠）
COLOR_TITLE_BG = "0078D4"      # 明るい青（タイトル）
COLOR_TOP_HEADER_BG = "1F4E79" # 濃紺（支店長 など最上位ラベル）
COLOR_SUBHEADER_BG = "5B9BD5"  # 中青（支店付 / 課長 / SA など）
COLOR_SECTION_BG = "DDEBF7"    # 淡い青（営 業 課 / 課サブセクション）
COLOR_TEXT_WHITE = "FFFFFF"
COLOR_TEXT_KANA = "555555"
COLOR_TEXT_YEAR = "555555"
COLOR_TEXT_KANJI = "111111"
COLOR_TEXT_DARK = "1F4E79"     # 淡い背景上のラベルテキスト
COLOR_TEXT_LEAVE = "888888"
COLOR_BORDER_THIN = "6E8FB7"   # 人物ブロック細枠
COLOR_BORDER_THICK = "1F4E79"  # ラベル左罫など

THIN_BORDER = Side(style="thin", color=COLOR_BORDER_THIN)
THICK_BORDER_DARK = Side(style="thick", color=COLOR_BORDER_THICK)


def _norm(s: str) -> str:
    """名前正規化（兼務マーカー除去・スペース除去）。"""
    s = re.sub(r"^[\s　]*兼\s*[)）)）]\s*", "", s or "")
    return s.replace("　", "").replace(" ", "").strip()


def _learn_reference_name(employee_service, emp, used_name: str) -> None:
    """マスター体制表でマッチした文字列を、その従業員の照合キーに追加する。

    既存の name / match_key と同じ場合は記録しない（冗長を防ぐ）。
    """
    used = (used_name or "").strip()
    if not used:
        return
    if used == (emp.name or ""):
        return
    existing_keys = [s.strip() for s in (emp.match_key or "").split(",") if s.strip()]
    if used in existing_keys:
        return
    existing_keys.append(used)
    new_value = ",".join(existing_keys)
    try:
        employee_service.update(emp.id, match_key=new_value)
    except Exception:
        pass


def _shorten_branch(name: str) -> str:
    """シート名として使用可能な形式に整形。

    マスター左側に書かれた支店名・部署名をそのまま使う方針。
    Excel のシート名禁止文字 (《 》 : \\ / ? * [ ]) のみ除去する。
    """
    s = name
    for ch in ("《", "》", ":", "\\", "/", "?", "*", "[", "]"):
        s = s.replace(ch, "")
    s = s.strip()
    # 31文字上限
    return s[:31] if len(s) > 31 else s


def _to_year4(raw) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    try:
        n = int(float(s))
    except ValueError:
        return s
    if n < 100:
        return str(1900 + n if n >= 50 else 2000 + n)
    return str(n)


def _setup_sheet(ws) -> None:
    """列幅: col A は左マージン (2.625)、col B 以降は等幅。

    派遣/契約社員などで人物が DATA_COL_END を超える列まで配置されることが
    あるので、設定範囲は十分広く取る（人物 13 スロット + バッファ）。
    """
    ws.column_dimensions["A"].width = MARGIN_COL_WIDTH
    # スロット 13 (col 3 + 13*6 = col 81) まで対応
    max_col = max(DATA_COL_END, 85)
    for c in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = UNIFORM_COL_WIDTH


def _setup_page(ws) -> None:
    """印刷時のページレイアウトを整える（A3横、余白小さめ、用紙中央寄せ）。"""
    from openpyxl.worksheet.page import PageMargins, PrintOptions
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A3  # A3 横（広い体制図向け）
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # 幅優先、高さは何ページにわたっても可
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.4, right=0.4, top=0.4, bottom=0.4,
        header=0.2, footer=0.2,
    )
    ws.print_options = PrintOptions(
        horizontalCentered=True,
        verticalCentered=False,
    )
    # 印刷範囲を全データ領域に設定
    ws.print_area = f"A1:{get_column_letter(DATA_COL_END)}{max(ws.max_row, 100)}"


def _set_block_row_heights(ws, row_start: int) -> None:
    """1人ブロック分の行高を設定（写真8行 + カナ/漢字/年の3行）。"""
    for r in range(row_start, row_start + PHOTO_ROWS):
        ws.row_dimensions[r].height = PHOTO_ROW_HEIGHT
    for k in range(LABEL_ROWS):
        ws.row_dimensions[row_start + PHOTO_ROWS + k].height = LABEL_ROW_HEIGHT


def _label_box(ws, row: int, col_start: int, col_end: int, text: str,
               *, size: int = 11, bold: bool = True,
               fill_color: str = COLOR_SECTION_BG,
               text_color: str = COLOR_TEXT_DARK,
               height: float | None = None,
               left_thick: bool = False):
    """ラベルボックス（背景色付きセル）を描画。"""
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font = Font(name=FONT_NAME, size=size, bold=bold, color=text_color)
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if left_thick:
        cell.border = Border(left=THICK_BORDER_DARK)
    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start,
                       end_row=row, end_column=col_end)
    if height:
        ws.row_dimensions[row].height = height


def _label_text(ws, row: int, col_start: int, col_end: int, text: str,
                *, size: int = 10, bold: bool = False,
                color: str = "111111", align: str = "center"):
    """3行ラベル（カナ/漢字/年）テキスト書き込み。"""
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font = Font(name=FONT_NAME, size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start,
                       end_row=row, end_column=col_end)


def _keep_side(side):
    """既存の Side が定義済みなら返す。なければ None。"""
    return side if (side is not None and side.style) else None


_PHOTO_INSET_EMU = 19050  # 約 2 px。0 だと写真縁がセル境界に一致して青枠を覆い隠す


def _insert_photo(ws, col0_1: int, row0_1: int, w_cols: int, h_rows: int, path: Path) -> None:
    """セル範囲にぴったり貼り付ける。img.width/height はあえてセットしない
    （TwoCellAnchor がストレッチを担うため、px指定すると逆効果になる）。

    `_PHOTO_INSET_EMU` 分だけ内側にオフセットして、テンプレートの青枠が
    写真下に隠れず4辺すべて見えるようにする。
    """
    img = XLImage(str(path))
    anchor = TwoCellAnchor(
        editAs="oneCell",
        _from=AnchorMarker(col=col0_1 - 1, colOff=_PHOTO_INSET_EMU,
                           row=row0_1 - 1, rowOff=_PHOTO_INSET_EMU),
        to=AnchorMarker(col=col0_1 - 1 + w_cols, colOff=-_PHOTO_INSET_EMU,
                        row=row0_1 - 1 + h_rows, rowOff=-_PHOTO_INSET_EMU),
    )
    img.anchor = anchor
    ws.add_image(img)


def _strip_ia_suffix(name: str) -> str:
    """氏名末尾の '(IA)' / '（IA）' を除去する。"""
    if not name:
        return name
    return name.replace("(IA)", "").replace("（IA）", "").strip()


def _is_ia_member(emp) -> bool:
    """その従業員が IA 区分かどうか。emp.is_ia フラグ または氏名に '(IA)' を含むかで判定。"""
    if not emp:
        return False
    if getattr(emp, "is_ia", False):
        return True
    name = emp.name or ""
    return "(IA)" in name or "（IA）" in name


def _person_display_name(person: PersonEntry, emp) -> str:
    """漢字氏名の表示。役職記号は基本除去するが「兼）」だけは復元する。
    末尾の (IA) は別セクションラベルに移すので氏名からは除去。
    """
    name = emp.name if emp else person.name
    name = _strip_ia_suffix(name)
    if "兼" in (person.marks or ""):
        return f"兼）{name}"
    return name


def _split_ia(persons: list[PersonEntry], lookup_fn) -> tuple[list[PersonEntry], list[PersonEntry]]:
    """jimu リストを IA と非IA に分割。"""
    ia_list: list[PersonEntry] = []
    non_ia: list[PersonEntry] = []
    for p in persons:
        emp, _, _, hidden = lookup_fn(p.name)
        if hidden:
            # 非表示者は両方からスキップ（呼び出し先で再度 lookup されてもhiddenで除外される）
            continue
        if _is_ia_member(emp):
            ia_list.append(p)
        else:
            non_ia.append(p)
    return ia_list, non_ia


def _jimu_label(branch_name: str) -> str:
    """事務女性の役職表示ラベル。支店=SA / 技術部・事業推進部=実務職 / その他=IA。"""
    if "支店" in branch_name:
        return "SA"
    if branch_name in ("技術部", "事業推進部"):
        return "実務職"
    return "IA"


def _normalize_year_display(s: str) -> str:
    """年次文字列から「再」を除去。"""
    return s.replace("再", "").strip()


def _draw_person_block(
    ws,
    col_start_1: int,
    row_start_1: int,
    person: PersonEntry,
    emp,
    photo_path: Path | None,
    is_on_leave: bool,
    show_year: bool = True,
) -> None:
    """1人分（写真8行 + カナ・漢字・年 の3行ラベル）を描画。

    人物スロット: col_start_1 から 5 列 (col_start_1+0 〜 col_start_1+4)、
    最右端 (+5) は次スロットとの間のギャップ。
    """
    _set_block_row_heights(ws, row_start_1)

    photo_col_end = col_start_1 + (PERSON_COL_W - 2)  # = +4 (5列の最右)

    # 写真+ラベルを細枠で囲む（show_year=False の場合、年行は枠外）
    border_rows = PHOTO_ROWS + (LABEL_ROWS if show_year else LABEL_ROWS - 1)
    for r in range(row_start_1, row_start_1 + border_rows):
        for c in range(col_start_1, photo_col_end + 1):
            is_top = (r == row_start_1)
            is_bottom = (r == row_start_1 + border_rows - 1)
            is_left = (c == col_start_1)
            is_right = (c == photo_col_end)
            ws.cell(row=r, column=c).border = Border(
                left=THIN_BORDER if is_left else None,
                right=THIN_BORDER if is_right else None,
                top=THIN_BORDER if is_top else None,
                bottom=THIN_BORDER if is_bottom else None,
            )

    img_to_use = photo_path if (photo_path and not is_on_leave) else PLACEHOLDER_PHOTO_PATH
    if img_to_use and img_to_use.is_file():
        _insert_photo(ws, col_start_1, row_start_1,
                      PERSON_COL_W - 1, PHOTO_ROWS, img_to_use)
    if is_on_leave:
        cell = ws.cell(row=row_start_1 + PHOTO_ROWS - 1, column=col_start_1,
                       value="休職中")
        cell.font = Font(name=FONT_NAME, size=9, italic=True, color=COLOR_TEXT_LEAVE)
        cell.alignment = Alignment(horizontal="center", vertical="bottom")

    label_row_kana = row_start_1 + PHOTO_ROWS
    label_row_kanji = label_row_kana + 1
    label_row_year = label_row_kanji + 1

    kana = (emp.name_kana if emp else None) or ""

    year_display: str = ""
    if emp and emp.join_year_text:
        year_display = str(emp.join_year_text)
    else:
        y4 = _to_year4(person.year)
        if y4:
            year_display = y4
        elif emp and emp.join_year:
            year_display = str(emp.join_year)

    # カナ: 全幅中央寄せ 10pt #555555
    _label_text(ws, label_row_kana, col_start_1, photo_col_end,
                kana, size=10, color=COLOR_TEXT_KANA, align="center")

    # 漢字: 全幅中央寄せ 11pt **非ボールド** #111111。
    # 役職記号は基本除去するが「兼）」だけは残す。
    _label_text(ws, label_row_kanji, col_start_1, photo_col_end,
                _person_display_name(person, emp),
                size=11, bold=False, color=COLOR_TEXT_KANJI, align="center")

    # 年: show_year=True かつ値がある時のみ書き込み。「再」は除去
    if show_year and year_display:
        cleaned_year = _normalize_year_display(year_display)
        if cleaned_year:
            _label_text(ws, label_row_year, col_start_1, photo_col_end,
                        cleaned_year, size=10, color=COLOR_TEXT_YEAR, align="center")


def _determine_leader_label(section_name: str, leader: PersonEntry) -> str:
    """先頭人物の役職ラベル（課長/課長代務/室長 等）を決定する。"""
    marks = leader.marks or ""
    is_acting = ("＊" in marks) or ("*" in marks)
    is_concurrent = "兼" in marks
    base = "室長" if ("室" in section_name and "課" not in section_name) else "課長"
    if is_acting:
        return f"{base}代務"
    if is_concurrent:
        return f"{base}（兼）"
    return base


def _section_title_format(section_name: str) -> str:
    """セクション名の表示形式。2〜3文字は文字間を空ける（例: '営業課' → '営 業 課'）。"""
    name = section_name.strip()
    if 2 <= len(name) <= 4:
        return " ".join(name)
    return name


def _slot_col(slot_index: int) -> int:
    """0始まりのスロット番号 → 開始列 (col_start_1)。"""
    return SHEET_START_COL + slot_index * PERSON_COL_W


def _slot_content_end(slot_col_start: int) -> int:
    return slot_col_start + (PERSON_COL_W - 2)  # 5列の右端


def _build_section(
    ws,
    row_start: int,
    section_name: str,
    course_rows: list[CourseRow],
    lookup_fn,
    per_course_jimu_label: str | None = None,
) -> tuple[int, list[int]]:
    """1セクションを描画。

    per_course_jimu_label が指定された場合、各課の写真+ラベル直下に
    その課に紐付く 実務職 (CourseRow.jimu + .keiyaku) を「実務職」(等)
    のラベル付きで描画する（技術部・事業推進部用）。
    """
    photo_row_starts: list[int] = []
    label_col_end = _slot_content_end(SHEET_START_COL)
    row = row_start

    for cr_idx, cr in enumerate(course_rows):
        is_continuation = (cr_idx > 0) and (not cr.course_name)

        if not is_continuation:
            # セクションヘッダ（営業課/設計課）は最初の課にだけ表示
            if cr_idx == 0:
                _label_box(ws, row, SHEET_START_COL, label_col_end,
                           _section_title_format(section_name),
                           size=12, bold=True,
                           fill_color=COLOR_SECTION_BG, text_color=COLOR_TEXT_DARK,
                           height=SECTION_HEADER_HEIGHT, left_thick=True)
                row += 1
                ws.row_dimensions[row].height = SPACER_HEIGHT
                row += 1

            if cr.course_name:
                _label_box(ws, row, SHEET_START_COL, label_col_end,
                           cr.course_name,
                           size=12, bold=True,
                           fill_color=COLOR_SECTION_BG, text_color=COLOR_TEXT_DARK,
                           height=SUBSECTION_HEIGHT, left_thick=True)
                row += 1
                ws.row_dimensions[row].height = SPACER_HEIGHT
                row += 1

            if cr.persons:
                role_lbl = _determine_leader_label(section_name, cr.persons[0])
                if role_lbl:
                    _label_box(ws, row, SHEET_START_COL, label_col_end,
                               role_lbl,
                               size=12, bold=True,
                               fill_color=COLOR_SUBHEADER_BG, text_color=COLOR_TEXT_WHITE,
                               height=ROLE_LABEL_HEIGHT, left_thick=True)
                    row += 1
                    ws.row_dimensions[row].height = SPACER_HEIGHT
                    row += 1

            base_slot = 0
        else:
            base_slot = 1

        photo_row_starts.append(row)

        shown_idx = 0
        for person in cr.persons:
            emp, photo_path, is_on_leave, hidden = lookup_fn(person.name)
            if hidden:
                continue
            col_start = _slot_col(base_slot + shown_idx)
            shown_idx += 1
            _draw_person_block(ws, col_start, row, person, emp, photo_path, is_on_leave)
        row += PHOTO_ROWS + LABEL_ROWS + 1

    # セクション内の全課行が終わったら、jimu/keiyaku/haken をセクション単位で集約して
    # ひとつの「実務職/IA」セクションとしてセクション末尾に配置する（部のみ）。
    if per_course_jimu_label:
        all_jimu = [p for cr in course_rows for p in cr.jimu]
        all_keiyaku = [p for cr in course_rows for p in cr.keiyaku]
        all_haken = [p for cr in course_rows for p in cr.haken]
        if all_jimu or all_keiyaku or all_haken:
            ia_persons, non_ia_jimu = _split_ia(all_jimu, lookup_fn)
            if ia_persons:
                row = _build_extra_persons(ws, row, "IA", ia_persons, lookup_fn)
            jimu_keiyaku = non_ia_jimu + all_keiyaku
            if jimu_keiyaku or all_haken:
                row = _build_extra_persons(ws, row, per_course_jimu_label,
                                            jimu_keiyaku, lookup_fn,
                                            haken=all_haken)

    return row, photo_row_starts


def _build_extra_persons(
    ws,
    row_start: int,
    title: str,
    persons: list[PersonEntry],
    lookup_fn,
    haken: list[PersonEntry] | None = None,
) -> int:
    """SA/実務職/IA などのサイドブロックを描画。年次は非表示。

    haken が与えられた場合、persons (jimu+keiyaku) を 1 段目、haken を 2 段目に
    強制改行して配置する（jimu+keiyaku の人数が少なくても改行）。
    """
    if not persons and not haken:
        return row_start

    label_col_start = _slot_col(1)
    label_col_end = _slot_content_end(label_col_start)
    _label_box(ws, row_start, label_col_start, label_col_end,
               title,
               size=12, bold=True,
               fill_color=COLOR_SUBHEADER_BG, text_color=COLOR_TEXT_WHITE,
               height=SECTION_HEADER_HEIGHT, left_thick=True)
    row = row_start + 1
    ws.row_dimensions[row].height = SPACER_HEIGHT
    row += 1

    per_row = 7
    base_slot = 1
    rendered_any = False

    def _render_group(group: list[PersonEntry], row: int) -> tuple[int, bool]:
        shown_idx = 0
        for person in group:
            emp, photo_path, is_on_leave, hidden = lookup_fn(person.name)
            if hidden:
                continue
            if shown_idx > 0 and shown_idx % per_row == 0:
                row += PERSON_TOTAL_H
            slot = base_slot + (shown_idx % per_row)
            col_start = _slot_col(slot)
            _draw_person_block(ws, col_start, row, person, emp, photo_path, is_on_leave,
                               show_year=False)
            shown_idx += 1
        return row, shown_idx > 0

    # 1段目: jimu+keiyaku
    if persons:
        row, did_render = _render_group(persons, row)
        if did_render:
            rendered_any = True
            if haken:
                row += PERSON_TOTAL_H   # 派遣は強制的に一段下

    # 2段目: haken
    if haken:
        row, did_render = _render_group(haken, row)
        if did_render:
            rendered_any = True

    if not rendered_any:
        return row_start
    return row + PERSON_TOTAL_H + 1


def build_workbook(
    branches: list[BranchData],
    employee_service: EmployeeService,
    photo_service: PhotoService,
) -> tuple[Workbook, dict]:
    stats = {"placed": 0, "unmatched": [], "on_leave": []}

    stats["ambiguous"] = []   # 曖昧マッチ（複数候補）

    stats["hidden_skipped"] = []

    def lookup(name: str):
        """戻り値: (emp, photo_path, is_on_leave, is_hidden)"""
        candidates_active, quality = employee_service.repo.find_by_text(
            name, only_active=True, return_quality=True
        )
        if candidates_active:
            if len(candidates_active) == 1:
                emp = candidates_active[0]
                if quality != "reference_name":
                    _learn_reference_name(employee_service, emp, name)
            else:
                emp = candidates_active[0]
                cand_names = ", ".join(c.name for c in candidates_active[:5])
                stats["ambiguous"].append(f"{name!r} → {len(candidates_active)}候補 [{cand_names}] (1人目採用)")
            if getattr(emp, "hidden", False):
                stats["hidden_skipped"].append(name)
                return None, None, False, True
            photo_path = photo_service.resolve(emp.photo_path)
            if photo_path:
                stats["placed"] += 1
            return emp, photo_path, False, False

        # 休職中含めて検索
        all_candidates = employee_service.repo.find_by_text(name, only_active=False)
        if all_candidates:
            emp = all_candidates[0]
            if getattr(emp, "hidden", False):
                stats["hidden_skipped"].append(name)
                return None, None, False, True
            stats["on_leave"].append(name)
            return emp, None, True, False

        stats["unmatched"].append(name)
        return None, None, False, False

    wb = Workbook()
    wb.remove(wb.active)

    for branch in branches:
        # 中身が空の支店/部署はシートを作らない
        total_persons = (
            len(branch.top_positions)
            + sum(len(cr.persons) for cs in branch.sections.values() for cr in cs)
            + len(branch.jimu) + len(branch.keiyaku) + len(branch.haken)
        )
        if total_persons == 0:
            continue

        sheet_name = _shorten_branch(branch.branch_name)[:31] or "Sheet"
        if sheet_name in wb.sheetnames:
            sheet_name = sheet_name + "_2"
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        _setup_sheet(ws)
        _setup_page(ws)

        # タイトル: 明るい青 #0078D4, 28pt Meiryo UI bold WHITE, 左寄せ indent=1, 高さ 50.25
        tcell = ws.cell(row=1, column=1, value=branch.branch_name)
        tcell.font = Font(name=FONT_NAME, bold=True, size=28, color=COLOR_TEXT_WHITE)
        tcell.fill = PatternFill("solid", fgColor=COLOR_TITLE_BG)
        tcell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=DATA_COL_END)
        ws.row_dimensions[1].height = TITLE_ROW_HEIGHT
        row = 3

        # 上位役職（支店長 = 濃紺、支店付/兼務 等 = 中青）
        # 兼務 はラベルを「支店付」にリネーム、人物 display に「兼）」を付与、
        # さらに視覚的に分離するため 1 スロット飛ばして配置する。
        if branch.top_positions:
            slotted_pairs: list[tuple[str, PersonEntry, int]] = []
            next_slot = 0
            for orig_label, raw_name in branch.top_positions:
                p = _parse_person(raw_name, None)
                if p is None:
                    continue
                if orig_label.strip() == "兼務":
                    display_label = "支店付"
                    if "兼" not in (p.marks or ""):
                        p.marks = "兼" + (p.marks or "")
                    # 兼務エントリは 1 スロット飛ばし
                    next_slot = max(next_slot + 1, 2)
                else:
                    display_label = orig_label
                slotted_pairs.append((display_label, p, next_slot))
                next_slot += 1

            for i, (lbl, _, slot) in enumerate(slotted_pairs):
                col_start = _slot_col(slot)
                col_end = _slot_content_end(col_start)
                if i == 0:
                    _label_box(ws, row, col_start, col_end, lbl,
                               size=12, bold=True,
                               fill_color=COLOR_TOP_HEADER_BG, text_color=COLOR_TEXT_WHITE,
                               height=TOP_HEADER_HEIGHT)
                else:
                    _label_box(ws, row, col_start, col_end, lbl,
                               size=10, bold=True,
                               fill_color=COLOR_SUBHEADER_BG, text_color=COLOR_TEXT_WHITE,
                               height=TOP_HEADER_HEIGHT)
            row += 1
            ws.row_dimensions[row].height = SPACER_HEIGHT
            row += 1
            for (_, p, slot) in slotted_pairs:
                emp, photo_path, is_on_leave, hidden = lookup(p.name)
                if hidden:
                    continue
                col_start = _slot_col(slot)
                _draw_person_block(ws, col_start, row, p, emp, photo_path, is_on_leave,
                                   show_year=False)
            row += PHOTO_ROWS + LABEL_ROWS + 1

        # 各セクション（●営業課/●設計課）
        # 部 (技術部・事業推進部) は per-課 で実務職セクションを差し込む
        is_dept = branch.branch_name in ("技術部", "事業推進部")
        per_course_label = _jimu_label(branch.branch_name) if is_dept else None
        section_photo_rows: dict[str, list[int]] = {}
        for section_name, course_rows in branch.sections.items():
            row, photo_starts = _build_section(ws, row, section_name, course_rows,
                                                lookup, per_course_jimu_label=per_course_label)
            section_photo_rows[section_name] = photo_starts

        # 派遣社員 → 設計課の最終課行に続けて配置（支店のみ）
        # 部 (技術部・事業推進部) は per-課 実務職セクション内で派遣も既に表示するため
        # branch-level の haken 配置は重複になるのでスキップ。
        extras = [] if is_dept else list(branch.haken)
        if extras:
            design_section_name = next(
                (sn for sn in section_photo_rows if "設計" in sn), None
            )
            target_row: int | None = None
            base_slot = 0
            occupied_slots = 0
            if design_section_name:
                starts = section_photo_rows[design_section_name]
                design_courses = branch.sections.get(design_section_name, [])
                if len(starts) >= 2:
                    target_row = starts[1]
                    if len(design_courses) >= 2:
                        if not design_courses[1].course_name:
                            base_slot = 1   # 継続行は slot1 開始
                        occupied_slots = len(design_courses[1].persons)
                elif starts:
                    target_row = starts[0]
                    occupied_slots = (len(design_courses[0].persons)
                                       if design_courses else 0)

            if target_row is not None:
                shown_idx = 0
                for person in extras:
                    emp, photo_path, is_on_leave, hidden = lookup(person.name)
                    if hidden:
                        continue
                    slot = base_slot + occupied_slots + shown_idx
                    shown_idx += 1
                    col_start = _slot_col(slot)
                    _draw_person_block(ws, col_start, target_row, person,
                                       emp, photo_path, is_on_leave)
            else:
                shown_idx = 0
                for person in extras:
                    emp, photo_path, is_on_leave, hidden = lookup(person.name)
                    if hidden:
                        continue
                    col_start = _slot_col(shown_idx)
                    shown_idx += 1
                    _draw_person_block(ws, col_start, row, person,
                                       emp, photo_path, is_on_leave)
                row += PHOTO_ROWS + LABEL_ROWS + 1

        # 支店の場合のみ、SA セクションを branch 全体で集約配置
        # （部の場合は各課直下に per-課 実務職を既に描画済み）
        # IA 区分の人は別途「IA」セクションに分離して先に描画する。
        if not is_dept:
            ia_persons, non_ia_jimu = _split_ia(branch.jimu, lookup)
            if ia_persons:
                row = _build_extra_persons(ws, row, "IA", ia_persons, lookup)
            sa_persons = non_ia_jimu + list(branch.keiyaku)
            if sa_persons:
                row = _build_extra_persons(ws, row,
                                            _jimu_label(branch.branch_name),
                                            sa_persons, lookup)

    return wb, stats


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("master", type=Path, help="入力 (202605.xlsx 形式のマスター)")
    parser.add_argument("output", type=Path, help="出力 xlsx")
    parser.add_argument("--data-dir", type=Path, default=None)
    args = parser.parse_args()

    if not args.master.is_file():
        print(f"[error] 入力ファイルが見つかりません: {args.master}", file=sys.stderr)
        return 1

    settings = Settings()
    data_dir = args.data_dir or settings.data_dir
    paths = DataPaths(data_dir)
    db = Database(paths.db_path)
    db.create_all()
    photos = PhotoService(paths.photos_dir)
    es = EmployeeService(db, photos)

    print(f"📂 入力: {args.master}")
    print(f"📦 DB  : {paths.db_path}")
    branches = parse_master(args.master)
    print(f"\n=== 解析結果 ===")
    for b in branches:
        total = sum(len(cr.persons) for cs in b.sections.values() for cr in cs)
        extra = len(b.jimu) + len(b.keiyaku) + len(b.haken)
        top = len(b.top_positions)
        print(f"  {b.branch_name}: 上位{top} / 総合職{total} / その他{extra}")

    wb, stats = build_workbook(branches, es, photos)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)

    print(f"\n=== 結果 ===")
    print(f"  写真挿入: {stats['placed']}枚")
    print(f"  未マッチ: {len(stats['unmatched'])}名")
    if stats['unmatched']:
        print(f"    {sorted(set(stats['unmatched']))[:15]}")
    print(f"  休職中  : {len(stats['on_leave'])}名")
    if stats.get('ambiguous'):
        print(f"\n⚠ 曖昧マッチ (複数候補から1人目採用): {len(stats['ambiguous'])}件")
        for line in stats['ambiguous'][:10]:
            print(f"   - {line}")
        print(f"   → 該当者の引用名(reference_name)を GUI で設定すると、確実にマッチします")
    print(f"\n✅ 出力完了: {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
