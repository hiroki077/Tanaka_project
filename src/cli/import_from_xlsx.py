"""完成版の体制表 xlsx から従業員情報＋顔写真を一括取り込み。

完成版（例: `2605顔写真体制.xlsx`）は支店別シートに以下が含まれる:
- 顔写真（埋め込み画像）
- カナ氏名（写真直下のセル）
- 入社年（カナ氏名の同行・右側）
- 漢字氏名（カナ氏名の次行）

このスクリプトは写真位置を起点に近傍セルを走査して人物データを抽出し、
DB に投入する。デフォルトはドライラン（プレビュー）、`--commit` で実投入。

実行例:
    python3 -m src.cli.import_from_xlsx file/2605顔写真体制.xlsx
    python3 -m src.cli.import_from_xlsx file/2605顔写真体制.xlsx --commit
"""
from __future__ import annotations
import argparse
import io
import re
import struct
import sys
import warnings
import zipfile
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

warnings.filterwarnings("ignore")

from openpyxl import load_workbook
from PIL import Image


# EMF レコードタイプ → 「offBmiSrc」フィールドのオフセット（レコード先頭から）
_EMR_DIB_OFFSETS = {
    76: 84,   # EMR_BITBLT
    77: 100,  # EMR_STRETCHBLT
    80: 60,   # EMR_SETDIBITSTODEVICE
    81: 48,   # EMR_STRETCHDIBITS
    114: 96,  # EMR_ALPHABLEND
}


def _extract_dibs_from_emf(emf: bytes) -> list[bytes]:
    """EMF をパースして埋め込み DIB を抽出し、BMPバイト列のリストを返す。

    EMF はベクター形式だが、内部に EMR_STRETCHDIBITS 等のレコードで
    ラスター画像 (DIB = BITMAPINFOHEADER + ピクセルデータ) を埋め込める。
    DIB に BITMAPFILEHEADER (14 bytes) を被せれば BMP として読める。
    """
    bmps: list[bytes] = []
    offset = 0
    while offset + 8 <= len(emf):
        try:
            iType, nSize = struct.unpack_from("<II", emf, offset)
        except struct.error:
            break
        if nSize < 8 or offset + nSize > len(emf):
            break
        if iType in _EMR_DIB_OFFSETS:
            off = _EMR_DIB_OFFSETS[iType]
            if off + 16 <= nSize:
                offBmi, cbBmi, offBits, cbBits = struct.unpack_from(
                    "<IIII", emf, offset + off
                )
                if (0 < offBmi < nSize and 0 < cbBmi < nSize and
                        0 < offBits < nSize and 0 < cbBits < nSize and
                        offBmi + cbBmi <= nSize and offBits + cbBits <= nSize):
                    bmi = emf[offset + offBmi: offset + offBmi + cbBmi]
                    bits = emf[offset + offBits: offset + offBits + cbBits]
                    bf_off = 14 + len(bmi)
                    bf_size = 14 + len(bmi) + len(bits)
                    bmp = (b"BM" + struct.pack("<I", bf_size) + b"\x00\x00\x00\x00" +
                           struct.pack("<I", bf_off) + bmi + bits)
                    bmps.append(bmp)
        offset += nSize
    return bmps


def _emf_to_jpeg(emf_bytes: bytes) -> bytes | None:
    """EMF を JPEG バイト列に変換。失敗時 None。"""
    for bmp in _extract_dibs_from_emf(emf_bytes):
        try:
            img = Image.open(io.BytesIO(bmp))
            img.load()
            if img.mode in ("RGBA", "LA", "P"):
                img = img.convert("RGB")
            out = io.BytesIO()
            img.save(out, format="JPEG", quality=92)
            return out.getvalue()
        except Exception:
            continue
    return None


@dataclass
class EmfImage:
    """drawing XML 経由で取得した EMF 画像（既に JPEG 化済み）。"""
    from_col_0: int
    from_row_0: int
    to_col_0: int
    to_row_0: int
    jpeg_bytes: bytes
    rotation_deg: int = 0


def _extract_emf_images(xlsx_path: Path) -> dict[str, list[EmfImage]]:
    """xlsx 内の全 EMF 画像を抽出し、シート名 → [EmfImage, ...] のマップを返す。

    openpyxl の ws._images は EMF/WMF をスキップするため、ここで補完する。
    """
    result: dict[str, list[EmfImage]] = defaultdict(list)
    with zipfile.ZipFile(xlsx_path) as z:
        wb_xml = z.read("xl/workbook.xml").decode("utf-8", errors="ignore")
        # シート名と r:id の対応
        sheet_entries = re.findall(
            r'<sheet[^>]+name="([^"]+)"[^>]+r:id="(rId\d+)"', wb_xml
        )
        wb_rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8", errors="ignore")
        rid_to_sheet_xml = dict(re.findall(
            r'Id="(rId\d+)"[^>]+Target="([^"]+)"', wb_rels
        ))

        for sheet_name, rid in sheet_entries:
            sheet_target = rid_to_sheet_xml.get(rid)
            if not sheet_target:
                continue
            sheet_path = f"xl/{sheet_target}" if not sheet_target.startswith("xl/") else sheet_target
            sheet_basename = sheet_path.split("/")[-1].replace(".xml", "")
            rels_path = f"xl/worksheets/_rels/{sheet_basename}.xml.rels"
            try:
                sheet_rels = z.read(rels_path).decode("utf-8", errors="ignore")
            except KeyError:
                continue
            m = re.search(r'Target="\.\./drawings/(drawing\d+\.xml)"', sheet_rels)
            if not m:
                continue
            drawing_name = m.group(1)
            try:
                drawing_xml = z.read(f"xl/drawings/{drawing_name}").decode("utf-8", errors="ignore")
                drawing_rels = z.read(f"xl/drawings/_rels/{drawing_name}.rels").decode("utf-8", errors="ignore")
            except KeyError:
                continue
            embed_to_media = dict(re.findall(
                r'Id="(rId\d+)"[^>]+Target="\.\./media/([^"]+)"', drawing_rels
            ))

            anchors = re.findall(
                r"<xdr:twoCellAnchor\b[^>]*>(.*?)</xdr:twoCellAnchor>",
                drawing_xml, re.DOTALL,
            )
            for a in anchors:
                fm = re.search(
                    r"<xdr:from>\s*<xdr:col>(\d+)</xdr:col>\s*<xdr:colOff>\d+</xdr:colOff>\s*"
                    r"<xdr:row>(\d+)</xdr:row>", a, re.DOTALL,
                )
                tm = re.search(
                    r"<xdr:to>\s*<xdr:col>(\d+)</xdr:col>\s*<xdr:colOff>\d+</xdr:colOff>\s*"
                    r"<xdr:row>(\d+)</xdr:row>", a, re.DOTALL,
                )
                embed_m = re.search(r'r:embed="(rId\d+)"', a)
                if not (fm and tm and embed_m):
                    continue
                media_name = embed_to_media.get(embed_m.group(1), "")
                if not media_name.lower().endswith(".emf"):
                    continue
                try:
                    emf_bytes = z.read(f"xl/media/{media_name}")
                except KeyError:
                    continue
                jpeg = _emf_to_jpeg(emf_bytes)
                if not jpeg:
                    continue
                rot_m = re.search(r'<a:xfrm\s+rot="(-?\d+)"', a)
                rot_deg = int(round((int(rot_m.group(1)) / 60000) % 360)) if rot_m else 0
                result[sheet_name].append(EmfImage(
                    from_col_0=int(fm.group(1)),
                    from_row_0=int(fm.group(2)),
                    to_col_0=int(tm.group(1)),
                    to_row_0=int(tm.group(2)),
                    jpeg_bytes=jpeg,
                    rotation_deg=rot_deg,
                ))
    return result

from ..config import Settings, DataPaths
from ..db import Database
from ..services import EmployeeService, PhotoService
from ..services.photo_service import normalize_image_bytes


# 抽出対象外シート（TOP・余白・スタッフ部署など、本部組織以外）
SKIP_SHEETS = {"Sheet1", "集合TOP"}

# 画像形式（写真として扱うもの）
PHOTO_FORMATS = {"jpeg", "jpg", "png"}

KATAKANA_RE = re.compile(r"^[ｦ-ﾟァ-ヶー\s　]+$")
YEAR_RE = re.compile(r"^(19|20)?\d{2}$")
YEAR_APOSTROPHE_RE = re.compile(r"^['’]?(\d{1,2})$")
WAREKI_RE = re.compile(r"^([HSRhsr平昭令])\s*(\d{1,2})$")
KANJI_RE = re.compile(r"[一-龥々ヶ]")

# 兼務マーカー（全角/半角の各種カッコバリエーション）
CONCURRENT_PREFIX_RE = re.compile(r"^[\s　]*兼\s*[)）)）]\s*")


_ANCHOR_BLOCK_RE = re.compile(
    r"<xdr:(?:twoCellAnchor|oneCellAnchor)\b[^>]*>(.*?)</xdr:(?:twoCellAnchor|oneCellAnchor)>",
    re.DOTALL,
)
_FROM_RE = re.compile(
    r"<xdr:from>\s*<xdr:col>(\d+)</xdr:col>\s*<xdr:colOff>\d+</xdr:colOff>\s*"
    r"<xdr:row>(\d+)</xdr:row>",
    re.DOTALL,
)
_ROT_RE = re.compile(r'<a:xfrm\s+rot="(-?\d+)"')


def extract_rotation_map(xlsx_path: Path) -> dict[tuple[int, int], int]:
    """xlsx 内の全 drawing XML を解析し、(from.col, from.row) → 回転度数 のマップを返す。

    Excel の回転は 1/60000 度単位で記録される（rot="5400000" → 90度）。
    位置 (col, row) は 0-based。openpyxl の anchor._from.col/.row と同じ系。
    """
    result: dict[tuple[int, int], int] = {}
    with zipfile.ZipFile(xlsx_path) as z:
        for name in z.namelist():
            if not (name.startswith("xl/drawings/") and name.endswith(".xml")):
                continue
            xml = z.read(name).decode("utf-8", errors="ignore")
            for block in _ANCHOR_BLOCK_RE.findall(xml):
                fm = _FROM_RE.search(block)
                if not fm:
                    continue
                col, row = int(fm.group(1)), int(fm.group(2))
                rm = _ROT_RE.search(block)
                if not rm:
                    continue
                deg = (int(rm.group(1)) / 60000) % 360
                if deg:
                    result[(col, row)] = int(round(deg))
    return result


def apply_rotation(data: bytes, degrees: int, fmt: str) -> bytes:
    """画像バイナリに指定角度（時計回り）の回転を適用する。"""
    img = Image.open(io.BytesIO(data))
    # PIL の rotate は反時計回りなので、Excelの時計回り→ -degrees
    rotated = img.rotate(-degrees, expand=True, resample=Image.BICUBIC)
    save_fmt = (fmt or img.format or "JPEG").upper()
    if save_fmt == "JPG":
        save_fmt = "JPEG"
    if save_fmt == "JPEG" and rotated.mode in ("RGBA", "LA", "P"):
        rotated = rotated.convert("RGB")
    out = io.BytesIO()
    if save_fmt == "JPEG":
        rotated.save(out, format=save_fmt, quality=92)
    else:
        rotated.save(out, format=save_fmt)
    return out.getvalue()


@dataclass
class ImportedEmployee:
    name: str
    name_kana: str | None
    join_year: int | None
    source_sheet: str   # 取得元シート名（DBには保存しない、ログ用）
    photo_bytes: bytes
    photo_format: str
    source_cell: str
    is_concurrent: bool = False  # 兼務フラグ

    @property
    def match_key(self) -> str:
        """既定の照合キーは漢字氏名の苗字部分（兼務マーカー除外、末尾空白除去）。"""
        if not self.name:
            return ""
        cleaned = CONCURRENT_PREFIX_RE.sub("", self.name)
        cleaned = cleaned.replace("　", " ").strip()
        head = cleaned.split(" ")[0] if " " in cleaned else cleaned
        return head.strip()

    @property
    def role(self) -> str | None:
        return "兼務" if self.is_concurrent else None


def _parse_year(value) -> int | None:
    """入社年のパース。以下の形式をサポート:

    - 4桁: `2004`, `1996`
    - 2桁: `04`, `96`（数値型・文字列型どちらも）
    - アポストロフィ付き: `'04`, `’04`
    - 和暦: `H17`, `S60`, `R5`, `平17`, `昭60`, `令5`
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        n = int(value)
        if 1900 <= n <= 2100:
            return n
        if 0 <= n < 100:
            return 1900 + n if n >= 50 else 2000 + n
        return None
    s = str(value).strip().replace("　", "")
    if not s:
        return None

    # 4桁 or 2桁
    m = YEAR_APOSTROPHE_RE.match(s)
    if m:
        n = int(m.group(1))
        if n < 100:
            return 1900 + n if n >= 50 else 2000 + n
        return n
    if YEAR_RE.match(s):
        n = int(s)
        if n < 100:
            return 1900 + n if n >= 50 else 2000 + n
        return n

    # 和暦
    m = WAREKI_RE.match(s)
    if m:
        era = m.group(1)
        n = int(m.group(2))
        if era in ("H", "h", "平"):
            return 1988 + n  # H1=1989
        if era in ("S", "s", "昭"):
            return 1925 + n  # S1=1926
        if era in ("R", "r", "令"):
            return 2018 + n  # R1=2019
    return None


def _is_katakana(value) -> bool:
    if value is None:
        return False
    s = str(value).strip()
    return bool(s) and bool(KATAKANA_RE.match(s))


def _has_kanji(value) -> bool:
    if value is None:
        return False
    return bool(KANJI_RE.search(str(value)))


_TEXT_KANJI_RE = re.compile(r"[一-龥々]")


def _scan_text_only_persons(ws, sheet_name: str,
                            existing_keys: set[tuple[str, str]]) -> list[ImportedEmployee]:
    """写真が無い「テキストだけの人物」をシートから拾う。

    パターン: 漢字氏名のセルがあり、すぐ上のセルにカタカナ氏名がある。
    既に画像経由で取れた人物 (existing_keys) は除外する。
    """
    found: list[ImportedEmployee] = []
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not v or not isinstance(v, str):
                continue
            text = v.strip()
            if not text or len(text) > 20:
                continue
            if not _TEXT_KANJI_RE.search(text):
                continue
            # ノイズ排除（セクションヘッダや役職、地名など）
            if any(k in text for k in [
                '《', '》', '●', '◆', '☆', '■', '◎',
                '兼務', '転入', '実務', '契約', '派遣', '休職', '産休',
                '部', '課', '店', '長', '本部', '事業', '管理', '推進',
                'スタッフ', '販売', '会議', 'GM', 'チーム', '受注', '責任',
                'INSTRUCTOR', '室',
            ]):
                continue
            # スペースで区切られているもの（漢字＋空白＋漢字）を氏名として優先
            # 上のセルにカタカナ（kana）があるかチェック
            row_n = cell.row
            col_n = cell.column
            kana_cell = ws.cell(row=row_n - 1, column=col_n).value
            if not kana_cell or not isinstance(kana_cell, str):
                continue
            kana = kana_cell.strip()
            if not (kana and KATAKANA_RE.match(kana)):
                continue

            # 兼務マーカー
            is_concurrent = bool(CONCURRENT_PREFIX_RE.match(text))
            clean_name = CONCURRENT_PREFIX_RE.sub("", text).strip()
            norm_name = clean_name.replace("　", "").replace(" ", "")
            norm_kana = kana.replace("　", "").replace(" ", "")
            key = (norm_name, norm_kana)
            if key in existing_keys:
                continue
            existing_keys.add(key)

            # 入社年（同行・右側）を試す
            year = None
            for c2 in range(col_n + 1, col_n + 8):
                v2 = ws.cell(row=row_n - 1, column=c2).value
                y = _parse_year(v2) if v2 is not None else None
                if y:
                    year = y
                    break

            from openpyxl.utils import get_column_letter
            cell_label = f"{get_column_letter(col_n)}{row_n}"
            found.append(ImportedEmployee(
                name=clean_name,
                name_kana=kana,
                join_year=year,
                source_sheet=sheet_name,
                photo_bytes=b"",
                photo_format="jpeg",
                source_cell=cell_label,
                is_concurrent=is_concurrent,
            ))
    return found


def _iter_photo_sources(ws, sheet_name: str, emf_map: dict[str, list[EmfImage]]):
    """ws._images (jpeg/png) と EMF 画像を統一インターフェースで yield する。

    yield: (from_col_0, from_row_0, to_col_0, to_row_0, photo_bytes, format_hint)
    """
    for img in ws._images:
        fmt = (getattr(img, "format", "") or "").lower()
        if fmt not in PHOTO_FORMATS:
            continue
        anchor = img.anchor
        if not (hasattr(anchor, "_from") and hasattr(anchor, "to") and anchor.to):
            continue
        try:
            data = img._data()
        except Exception:
            data = b""
        yield (anchor._from.col, anchor._from.row,
               anchor.to.col, anchor.to.row, data, fmt, False)

    for emf in emf_map.get(sheet_name, []):
        yield (emf.from_col_0, emf.from_row_0, emf.to_col_0, emf.to_row_0,
               emf.jpeg_bytes, "jpeg", True)


def extract_from_sheet(
    ws,
    sheet_name: str,
    rotation_map: dict[tuple[int, int], int],
    emf_map: dict[str, list[EmfImage]] | None = None,
) -> list[ImportedEmployee]:
    """1シートから写真＋人物情報を抽出。Excel上の回転は物理回転として適用する。

    ws._images に加え、EMF 画像（openpyxl がサポートしない形式）も拾う。
    """
    if emf_map is None:
        emf_map = {}
    results: list[ImportedEmployee] = []
    seen_keys: set[tuple[str, str]] = set()

    for (from_col_0, from_row_0, to_col_0, to_row_0,
         photo_bytes_src, fmt, is_emf_origin) in _iter_photo_sources(ws, sheet_name, emf_map):
        rotation_deg = rotation_map.get((from_col_0, from_row_0), 0)

        # 探索範囲を写真より左右1列ずつ拡張（氏名セルが写真の端より1列外側にあるケース対応）
        # 例: 写真 col=17..21 (R-V) のとき、氏名は W27 (col=22) にあることがある

        # 1-based の探索範囲（画像直下のラベル枠）。氏名が写真の端を1列はみ出る
        # ケースに対応するため、左右1列ずつ余分に探索する。
        col_lo = max(1, from_col_0)
        col_hi = to_col_0 + 2
        row_start = to_row_0 + 1

        kana = None
        year = None
        kanji = None
        anchor_col = None
        kana_row = None
        for offset in range(0, 5):
            row = row_start + offset
            for col in range(col_lo, col_hi + 1):
                v = ws.cell(row=row, column=col).value
                if _is_katakana(v):
                    kana = str(v).strip()
                    anchor_col = col
                    kana_row = row
                    break
            if kana:
                break

        if kana and anchor_col is not None and kana_row is not None:
            # 同行の右側で年を探す
            for c2 in range(anchor_col + 1, anchor_col + 8):
                v = ws.cell(row=kana_row, column=c2).value
                y = _parse_year(v) if v is not None else None
                if y:
                    year = y
                    break
            # 次行で漢字氏名を探す
            for r2 in (kana_row + 1, kana_row + 2):
                v = ws.cell(row=r2, column=anchor_col).value
                if v and _has_kanji(v):
                    kanji = str(v).strip()
                    break

        if not kanji:
            # 漢字氏名が拾えない場合スキップ（写真だけ→人物特定不可）
            continue

        photo_bytes = photo_bytes_src

        if rotation_deg and photo_bytes:
            try:
                photo_bytes = apply_rotation(photo_bytes, rotation_deg, fmt)
            except Exception as e:
                print(f"  [warn] rotation failed at {sheet_name} "
                      f"({from_col_0},{from_row_0}) +{rotation_deg}°: {e}",
                      file=sys.stderr)

        from openpyxl.utils import get_column_letter
        cell_label = f"{get_column_letter(anchor_col or col_lo)}{kana_row or row_start}"

        is_concurrent = bool(CONCURRENT_PREFIX_RE.match(kanji))
        clean_name = CONCURRENT_PREFIX_RE.sub("", kanji).strip()

        norm_n = clean_name.replace("　", "").replace(" ", "")
        norm_k = (kana or "").replace("　", "").replace(" ", "")
        key = (norm_n, norm_k)
        if key in seen_keys:
            continue
        seen_keys.add(key)

        results.append(ImportedEmployee(
            name=clean_name,
            name_kana=kana.strip() if kana else None,
            join_year=year,
            source_sheet=sheet_name,
            photo_bytes=photo_bytes,
            photo_format=fmt,
            source_cell=cell_label,
            is_concurrent=is_concurrent,
        ))

    # 写真がない「テキストのみ」の人物も拾う
    results.extend(_scan_text_only_persons(ws, sheet_name, seen_keys))
    return results


def import_workbook(path: Path) -> list[ImportedEmployee]:
    wb = load_workbook(path)
    rotation_map = extract_rotation_map(path)
    if rotation_map:
        print(f"  検出した回転メタデータ: {len(rotation_map)}件")
    emf_map = _extract_emf_images(path)
    if emf_map:
        total_emf = sum(len(v) for v in emf_map.values())
        print(f"  EMF 写真も検出: {total_emf}枚（jpeg/png 以外の埋め込み）")

    # ワークブック全体での重複排除キー（正規化氏名 + 正規化カナ）
    global_keys: dict[tuple[str, str], ImportedEmployee] = {}
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        records = extract_from_sheet(ws, sheet_name, rotation_map, emf_map)
        for rec in records:
            norm_n = (rec.name or "").replace("　", "").replace(" ", "")
            norm_k = (rec.name_kana or "").replace("　", "").replace(" ", "")
            key = (norm_n, norm_k)
            existing = global_keys.get(key)
            if existing is None:
                global_keys[key] = rec
            else:
                # 重複時の優先順位: 写真ありを優先、それ以外は最初を残す
                if not existing.photo_bytes and rec.photo_bytes:
                    global_keys[key] = rec
    return list(global_keys.values())


def commit_to_db(records: list[ImportedEmployee], paths: DataPaths) -> int:
    """既存DBをクリーンにしてから投入（再投入を想定）。"""
    import uuid

    db = Database(paths.db_path)
    db.create_all()
    photos = PhotoService(paths.photos_dir)
    service = EmployeeService(db, photos)

    # 既存従業員を全削除（写真ファイルもクリーンアップ）
    existing = service.list()
    for emp in existing:
        service.delete(emp.id)

    inserted = 0
    for rec in records:
        photo_name = None
        if rec.photo_bytes:
            normalized, ext = normalize_image_bytes(rec.photo_bytes, format_hint=rec.photo_format)
            photo_name = f"{uuid.uuid4().hex}.{ext}"
            (paths.photos_dir / photo_name).write_bytes(normalized)

        emp = service.create(
            name=rec.name,
            name_kana=rec.name_kana,
            match_key=rec.match_key,
            join_year=rec.join_year,
            role=rec.role,
            status="在職",
        )
        if photo_name:
            service.update(emp.id, photo_path=photo_name)
        inserted += 1
    return inserted


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx", type=Path, help="取り込み元 xlsx ファイル")
    parser.add_argument("--commit", action="store_true",
                        help="DBに実投入する（指定なしはドライラン）")
    parser.add_argument("--data-dir", type=Path, default=None,
                        help="データフォルダ（既定: settings.json の data_dir）")
    args = parser.parse_args()

    if not args.xlsx.is_file():
        print(f"[error] ファイルが見つかりません: {args.xlsx}", file=sys.stderr)
        return 1

    print(f"📂 解析中: {args.xlsx}")
    records = import_workbook(args.xlsx)

    by_sheet: dict[str, list[ImportedEmployee]] = {}
    for r in records:
        by_sheet.setdefault(r.source_sheet, []).append(r)

    print(f"\n=== 抽出結果: 合計 {len(records)} 名 ===")
    for sheet, items in by_sheet.items():
        print(f"\n[{sheet}] {len(items)}名")
        for r in items[:5]:
            year = r.join_year if r.join_year else "?"
            print(f"  - {r.name} ({r.name_kana}, {year}入社) "
                  f"key={r.match_key} cell={r.source_cell}")
        if len(items) > 5:
            print(f"  ... 他 {len(items) - 5}名")

    if not args.commit:
        print("\n[dry-run] --commit を付けるとDBに投入します。")
        return 0

    settings = Settings()
    data_dir = args.data_dir or settings.data_dir
    paths = DataPaths(data_dir)
    print(f"\n📦 DB投入先: {paths.data_dir}")
    print(f"   DB: {paths.db_path}")
    print(f"   写真: {paths.photos_dir}")
    inserted = commit_to_db(records, paths)
    print(f"\n✅ {inserted}名を投入しました。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
