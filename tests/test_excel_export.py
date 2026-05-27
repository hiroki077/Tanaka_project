"""Excel出力エンジンのスモークテスト。

実行方法:
    cd Miki_application
    python -m tests.test_excel_export
"""
from __future__ import annotations
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook, load_workbook
from PIL import Image as PILImage

from src.db import Database
from src.config import DataPaths
from src.services import (
    EmployeeService,
    PhotoService,
    ExcelExportService,
    ExportOptions,
    first_candidate_resolver,
)


def _make_dummy_photo(path: Path, color: tuple[int, int, int]) -> None:
    img = PILImage.new("RGB", (200, 250), color=color)
    img.save(path, format="JPEG")


def _make_template_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "東京東"
    ws["B3"] = "{{photo:藤堂}}"
    ws["B12"] = "藤堂"
    ws["E12"] = "綾部"
    ws["H12"] = "佐藤"  # 同名複数 → ambiguity_resolver で解決
    ws["K12"] = "青木"   # 休職中
    ws["N3"] = "{{photo:佐藤#id=__SATO_N_ID__}}"  # ID指定はテスト内で置換
    ws["N12"] = "佐藤"

    wb.save(path)


def main() -> int:
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        paths = DataPaths(tmp / "data")

        # DBセットアップ
        db = Database(paths.db_path)
        db.create_all()
        photos = PhotoService(paths.photos_dir)
        es = EmployeeService(db, photos)

        # 写真ファイル作成
        p_todo = tmp / "todo.jpg"
        p_ayabe = tmp / "ayabe.jpg"
        p_sato_n = tmp / "sato_north.jpg"
        p_sato_s = tmp / "sato_south.jpg"
        _make_dummy_photo(p_todo, (200, 100, 100))
        _make_dummy_photo(p_ayabe, (100, 200, 100))
        _make_dummy_photo(p_sato_n, (100, 100, 200))
        _make_dummy_photo(p_sato_s, (200, 200, 100))

        # 従業員登録（所属情報はDBに持たない）
        todo = es.create(name="藤堂 誠司", match_key="藤堂")
        es.set_photo(todo.id, str(p_todo))
        ayabe = es.create(name="綾部 良一", match_key="綾部")
        es.set_photo(ayabe.id, str(p_ayabe))
        sato_n = es.create(name="佐藤 北", match_key="佐藤")
        es.set_photo(sato_n.id, str(p_sato_n))
        sato_s = es.create(name="佐藤 南", match_key="佐藤")
        es.set_photo(sato_s.id, str(p_sato_s))
        es.create(name="青木 太郎", match_key="青木", status="休職中")

        template = tmp / "template.xlsx"
        _make_template_xlsx(template)
        # ID指定プレースホルダーに佐藤南のIDを埋め込む
        import openpyxl
        wb_tpl = openpyxl.load_workbook(template)
        ws_tpl = wb_tpl.active
        ws_tpl["N3"] = ws_tpl["N3"].value.replace("__SATO_N_ID__", str(sato_s.id))
        wb_tpl.save(template)

        out = tmp / "output.xlsx"
        svc = ExcelExportService(db, es, photos, first_candidate_resolver)
        result = svc.export(ExportOptions(
            template_path=template,
            output_path=out,
        ))

        print(f"sheets_processed: {result.sheets_processed}")
        print(f"photos_inserted : {result.photos_inserted}")
        print(f"skipped_no_photo: {result.skipped_no_photo}")
        print(f"skipped_on_leave: {result.skipped_on_leave}")
        print(f"unmatched       : {result.unmatched}")
        print(f"warnings        : {result.warnings}")

        # 検証
        assert out.is_file(), "出力ファイルが作成されていない"
        wb = load_workbook(out)
        ws_east = wb["東京東"]
        n_east = len(ws_east._images)
        print(f"東京東 画像数: {n_east}")

        # 期待:
        #   B3 {{photo:藤堂}}     → 1
        #   B12 藤堂(氏名検出)    → 1
        #   E12 綾部              → 1
        #   H12 佐藤(同名2)       → resolver で佐藤北選択 → 1
        #   N3 {{photo:#id=...}} → 佐藤南指定 → 1
        #   N12 佐藤(同名2)       → 上で学習済みオーバーライドが効くがセル違うので resolver 再走 → 佐藤北 → 1
        #   K12 青木(休職)        → スキップ
        assert n_east == 6, f"期待6枚 実際{n_east}枚"
        assert any("青木" in s for s in result.skipped_on_leave), "休職中スキップが記録されていない"

        print("\n[OK] テスト成功")
        return 0


if __name__ == "__main__":
    sys.exit(main())
